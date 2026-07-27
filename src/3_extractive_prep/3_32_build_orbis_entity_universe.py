"""
Build the full Orbis extractive-entity universe used by 1_7 to attach an HQ country
to every EITI company.

The earlier matcher only saw `orbis_broad_extractive_companies.csv` (~6k *deduplicated
top-level* companies).  But the raw Orbis pull (`data/raw/orbis/`) contains the full
extractive universe — ~600k entities, **each carrying its GUO bvd-id / name / type /
country directly** — i.e. it already includes the local operating subsidiaries
("Total E&P Angola", "Shell Nigeria", "Kibali Goldmines SA", …).  Matching an EITI
company string against *that* universe (blocked by the country it pays in) and reading
off the matched entity's GUO country is the proper way to get the HQ.

Sources (all in data/raw/orbis/):
  extractives_guo_orbis*.csv / .xlsx   — Company name, BvD ID, GUO BvD/Name/Type/Country  (entity→GUO)
  extractives_static_orbis*.csv         — BvD ID, Country ISO, NACE core, No of subsidiaries (ultimately-owned included)
  extractives_financials_orbis*.csv     — BvD ID, Operating revenue (Turnover) th USD 2016..2025   (NOTE: batches 33/35/36/38/39/42 missing → ~15% of entities have no revenue)

Output: data/intermediate/extractive/orbis_entity_universe.csv
  one row per Orbis bvd_id: name, entity_country, nace_core, n_subsidiaries,
  guo_bvd_id, guo_name, guo_type, guo_country, peak_operating_revenue_usd, in_cbcr_universe
  where in_cbcr_universe = (n_subsidiaries >= 2)  OR  (peak revenue >= €750M)   — a multi-entity
  extractive group above the CbCR turnover threshold; revenue missing ⇒ not disqualifying.
"""
import csv
import glob
import re
import sys
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from _paths import EXT_INT

RAW_ORBIS = (EXT_INT / ".." / ".." / "raw" / "orbis").resolve()
OUT = EXT_INT / "orbis_entity_universe.csv"
CBCR_REVENUE_THRESHOLD_TH_USD = 750_000.0   # €750M ≈ $750M, Orbis revenue is in thousands USD
MIN_SUBS_FOR_MNE = 2


def _norm_hdr(s):
    """Normalise a header cell for tolerant matching across Orbis web-export and
    flatfile naming: lowercase, underscores→spaces, collapse whitespace, and
    strip the GUO ownership-threshold token so `GUO 50 - Name` / `GUO_50_Name`
    both match the `guo - name` prefix used below."""
    s = str(s).strip().lower().replace("_", " ")
    s = re.sub(r"\bguo\s*\d+(?:\.\d+)?\b", "guo", s)   # "guo 50 -" / "guo 25.01 -" -> "guo -"
    s = re.sub(r"\s+", " ", s)
    return s


def _hdr_idx(hdr, *prefixes):
    for i, c in enumerate(hdr):
        cc = _norm_hdr(c)
        for p in prefixes:
            if cc.startswith(_norm_hdr(p)):
                return i
    return None


def _num(x):
    if x is None:
        return None
    x = str(x).strip().replace(",", "").replace(" ", "")
    if x in ("", "n.a.", "n.s.", "na", "N/A", "-"):
        return None
    try:
        return float(x)
    except ValueError:
        return None


def read_guo():
    """bvd_id -> (name, guo_bvd, guo_name, guo_type, guo_country)"""
    out = {}
    for f in sorted(glob.glob(str(RAW_ORBIS / "extractives_guo_orbis*.csv"))):
        with open(f, encoding="utf-8-sig", newline="") as fh:
            r = csv.reader(fh)
            hdr = next(r)
            i_name = _hdr_idx(hdr, "company name")
            i_bvd = _hdr_idx(hdr, "bvd id")
            i_gb = _hdr_idx(hdr, "guo - bvd id", "guo  - bvd id")
            i_gn = _hdr_idx(hdr, "guo - name")
            i_gt = _hdr_idx(hdr, "guo - type")
            i_gc = _hdr_idx(hdr, "guo - country")
            for row in r:
                if i_bvd is None or i_bvd >= len(row):
                    continue
                bvd = row[i_bvd].strip()
                if not bvd:
                    continue
                out[bvd] = (
                    row[i_name].strip() if i_name is not None and i_name < len(row) else "",
                    row[i_gb].strip() if i_gb is not None and i_gb < len(row) else "",
                    row[i_gn].strip() if i_gn is not None and i_gn < len(row) else "",
                    row[i_gt].strip() if i_gt is not None and i_gt < len(row) else "",
                    row[i_gc].strip() if i_gc is not None and i_gc < len(row) else "",
                )
    # the two batches stuck in .xlsx
    for f in sorted(glob.glob(str(RAW_ORBIS / "extractives_guo_orbis*.xlsx"))):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            hdr = [str(c) if c is not None else "" for c in next(rows)]
            i_name = _hdr_idx(hdr, "company name"); i_bvd = _hdr_idx(hdr, "bvd id")
            i_gb = _hdr_idx(hdr, "guo - bvd id", "guo  - bvd id"); i_gn = _hdr_idx(hdr, "guo - name")
            i_gt = _hdr_idx(hdr, "guo - type"); i_gc = _hdr_idx(hdr, "guo - country")
            for row in rows:
                if i_bvd is None or i_bvd >= len(row) or not row[i_bvd]:
                    continue
                bvd = str(row[i_bvd]).strip()
                out.setdefault(bvd, (
                    str(row[i_name]).strip() if i_name is not None and i_name < len(row) and row[i_name] else "",
                    str(row[i_gb]).strip() if i_gb is not None and i_gb < len(row) and row[i_gb] else "",
                    str(row[i_gn]).strip() if i_gn is not None and i_gn < len(row) and row[i_gn] else "",
                    str(row[i_gt]).strip() if i_gt is not None and i_gt < len(row) and row[i_gt] else "",
                    str(row[i_gc]).strip() if i_gc is not None and i_gc < len(row) and row[i_gc] else "",
                ))
            wb.close()
        except Exception as e:
            print(f"  [warn] could not read {Path(f).name}: {e}")
    return out


def read_static():
    """bvd_id -> (entity_country, nace_core, n_subsidiaries)"""
    out = {}
    for f in sorted(glob.glob(str(RAW_ORBIS / "extractives_static_orbis*.csv"))):
        with open(f, encoding="utf-8-sig", newline="") as fh:
            r = csv.reader(fh)
            hdr = next(r)
            i_bvd = _hdr_idx(hdr, "bvd id")
            i_cc = _hdr_idx(hdr, "country iso code")
            i_nace = _hdr_idx(hdr, "nace rev. 2, core code")
            i_subs = _hdr_idx(hdr, "no of subsidiaries")
            for row in r:
                if i_bvd is None or i_bvd >= len(row):
                    continue
                bvd = row[i_bvd].strip()
                if not bvd:           # continuation row (secondary NACE) — skip
                    continue
                out[bvd] = (
                    (row[i_cc].strip() if i_cc is not None and i_cc < len(row) else ""),
                    (row[i_nace].strip() if i_nace is not None and i_nace < len(row) else ""),
                    _num(row[i_subs]) if i_subs is not None and i_subs < len(row) else None,
                )
    return out


def read_financials():
    """bvd_id -> peak operating revenue (th USD) over 2016..2025"""
    out = {}
    for f in sorted(glob.glob(str(RAW_ORBIS / "extractives_financials_orbis*.csv"))):
        with open(f, encoding="utf-8-sig", newline="") as fh:
            r = csv.reader(fh)
            hdr = next(r)
            i_bvd = _hdr_idx(hdr, "bvd id")
            rev_cols = [i for i, c in enumerate(hdr) if c.strip().lower().startswith("operating revenue")]
            for row in r:
                if i_bvd is None or i_bvd >= len(row):
                    continue
                bvd = row[i_bvd].strip()
                if not bvd:
                    continue
                vals = [_num(row[i]) for i in rev_cols if i < len(row)]
                vals = [v for v in vals if v is not None]
                if vals:
                    out[bvd] = max(out.get(bvd, 0.0), max(vals))
    return out


def main():
    print("Reading raw Orbis extracts ...")
    guo = read_guo()
    print(f"  guo:        {len(guo):,} entities")
    static = read_static()
    print(f"  static:     {len(static):,} entities")
    fin = read_financials()
    print(f"  financials: {len(fin):,} entities with a revenue figure")

    all_bvd = set(guo) | set(static)
    print(f"  union:      {len(all_bvd):,} bvd_ids")

    cols = ["bvd_id", "name", "entity_country", "nace_core", "n_subsidiaries",
            "guo_bvd_id", "guo_name", "guo_type", "guo_country", "peak_operating_revenue_th_usd", "in_cbcr_universe"]
    n_mne = 0
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(cols)
        for bvd in sorted(all_bvd):
            gname, gb, gn, gt, gc = guo.get(bvd, ("", "", "", "", ""))
            ec, nace, nsub = static.get(bvd, ("", "", None))
            rev = fin.get(bvd)
            name = gname or ""
            # the GUO of a top-level company is itself; if guo missing, fall back to entity country as guo country
            guo_country = gc or ec
            in_mne = ((nsub is not None and nsub >= MIN_SUBS_FOR_MNE) or
                      (rev is not None and rev >= CBCR_REVENUE_THRESHOLD_TH_USD))
            if in_mne:
                n_mne += 1
            w.writerow([bvd, name, ec, nace, ("" if nsub is None else int(nsub)),
                        gb, gn, gt, guo_country, ("" if rev is None else f"{rev:.0f}"), int(in_mne)])
    print(f"\nWrote {OUT}  ({len(all_bvd):,} entities; {n_mne:,} flagged in_cbcr_universe = {100*n_mne/max(1,len(all_bvd)):.0f}%)")
    # quick coverage notes
    no_guo_country = sum(1 for bvd in all_bvd if not (guo.get(bvd, ("", "", "", "", ""))[4] or static.get(bvd, ("", "", None))[0]))
    print(f"  entities with no GUO country and no entity country: {no_guo_country:,}")
    n_no_rev = len(all_bvd) - len(set(fin) & all_bvd)
    print(f"  entities with no revenue figure: {n_no_rev:,} ({100*n_no_rev/len(all_bvd):.0f}%)  ← financials batches 33/35/36/38/39/42 are missing from the pull")


if __name__ == "__main__":
    main()
