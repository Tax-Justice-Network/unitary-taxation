# %%
"""
3_12 — Parse EITI production volumes.

Parse the EITI Summary Data XLSX templates into a long-format production +
exports table (one row per iso3 x year x commodity x measure).

Template variants supported:
  - v2 / "Summary Data v2.xlsx"     : sheet 'Part 2 - Disclosure checklist' has commodity rows
  - 2.0 / "Summary Data 2.0.xlsx"   : same v2 layout, slightly older labels
  - _san / "Summary Data_san.xlsx"  : older template, structure varies (best-effort fallback)

Production data lives in 'Part 2 - Disclosure checklist' under section header
"EITI Requirement 3.2: Production by commodity" (or "Production volume and value").
Exports under "EITI Requirement 3.3" / "Export volume and value".

Output columns:
  iso3, fy_label, fy_start_year, fy_end_year, section, commodity_name, hs_code,
  measure (volume|value), value, unit, source_text, source_file

Pipeline: Extractive prep, stage 3_12 — ingests the EITI production-volume
product (the second EITI ingest); feeds the 3_2x rent-combination stage.

Reads:
  data/raw/extractive/eiti_reports/<Country>/FY<years> <Country> Summary Data*.xlsx  — EITI Summary Data templates

Writes:
  data/intermediate/extractive/eiti_xlsx_production.csv  — iso3 x year x commodity x measure

Usage:
  python 3_12_parse_eiti_production_volumes.py

Author: Alison Schultz.
Last updated: 2026-07-25.
"""

# %% MARK: 1. Setup
import csv
import re
import sys
from pathlib import Path

import openpyxl
import warnings

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from _paths import RAW, EXT_INT

# %% MARK: 2. Config and lookups
EITI_ROOT = RAW / "extractive/eiti_reports"                    # external EITI XLSX templates
OUT_PATH = EXT_INT / "eiti_xlsx_production.csv"     # parsed → intermediate

# Country folder name -> ISO3. Drive folder uses common names.
NAME_TO_ISO3 = {
    "Afghanistan": "AFG", "Albania": "ALB", "Angola": "AGO",
    "Argentina": "ARG", "Armenia": "ARM",
    "Azerbaijan": "AZE", "Burkina Faso": "BFA", "Cameroon": "CMR",
    "Central African Republic": "CAF", "Chad": "TCD", "Colombia": "COL",
    "Cote d'Ivoire": "CIV", "Cote dIvoire": "CIV", "Côte d'Ivoire": "CIV",
    "Democratic Republic of the Congo": "COD", "Dominican Republic": "DOM",
    "Ecuador": "ECU", "Ethiopia": "ETH", "Gabon": "GAB", "Germany": "DEU",
    "Ghana": "GHA", "Guatemala": "GTM", "Guinea": "GIN", "Guyana": "GUY",
    "Honduras": "HND", "Indonesia": "IDN", "Iraq": "IRQ", "Kazakhstan": "KAZ",
    "Kyrgyzstan": "KGZ", "Liberia": "LBR", "Madagascar": "MDG", "Malawi": "MWI",
    "Mali": "MLI", "Mauritania": "MRT", "Mexico": "MEX", "Mongolia": "MNG",
    "Mozambique": "MOZ", "Myanmar": "MMR", "Netherlands": "NLD", "Niger": "NER",
    "Nigeria": "NGA", "Norway": "NOR", "Papua New Guinea": "PNG", "Peru": "PER",
    "Philippines": "PHL", "Republic of Congo": "COG",
    "Republic of the Congo": "COG", "Sao Tome and Principe": "STP",
    "São Tomé and Príncipe": "STP", "Senegal": "SEN", "Sierra Leone": "SLE",
    "Solomon Islands": "SLB", "Suriname": "SUR", "Tajikistan": "TJK",
    "Tanzania": "TZA",
    "Timor-Leste": "TLS", "Togo": "TGO", "Trinidad and Tobago": "TTO",
    "Tunisia": "TUN", "Uganda": "UGA", "Ukraine": "UKR",
    "United Kingdom": "GBR", "United States": "USA", "Zambia": "ZMB",
}

# Commodity-name (with HS code) -> our pipeline category. The pipeline groups
# carve-out base into oil_gas / coal / minerals; "other" is captured but not
# used in the carve-out.
HS_TO_CATEGORY = {
    "2709": "oil_gas",     # Crude oil
    "2710": "oil_gas",     # Petroleum products (refined)
    "2711": "oil_gas",     # Natural gas
    "2701": "coal",        # Coal
    "2702": "coal",        # Lignite
    "2603": "minerals",    # Copper ores
    "2601": "minerals",    # Iron ores
    "2602": "minerals",    # Manganese ores
    "2604": "minerals",    # Nickel ores
    "2605": "minerals",    # Cobalt ores
    "2606": "minerals",    # Aluminium ores (bauxite)
    "2607": "minerals",    # Lead ores
    "2608": "minerals",    # Zinc ores
    "2609": "minerals",    # Tin ores
    "2610": "minerals",    # Chromium ores
    "2611": "minerals",    # Tungsten ores
    "2613": "minerals",    # Molybdenum ores
    "2614": "minerals",    # Titanium ores
    "2615": "minerals",    # Niobium/tantalum/vanadium/zirconium
    "2616": "minerals",    # Precious metal ores
    "2617": "minerals",    # Other ores
    "7102": "minerals",    # Diamonds
    "7106": "minerals",    # Silver
    "7108": "minerals",    # Gold
    "7110": "minerals",    # Platinum-group metals
    "2510": "minerals",    # Phosphates
    # Construction/aggregate categories below: tracked but NOT in carve-out
    "2501": "other",       # Salt
    "2505": "other",       # Sand
    "2515": "other",       # Marble (decorative)
    "2516": "other",       # Granite
}

# Section header substrings (case-insensitive) used to identify Production vs
# Exports tables in the disclosure checklist sheet.
PRODUCTION_HEADERS = (
    "production by commodity",
    "production volume and value",
    "requirement 3.2",
)
EXPORTS_HEADERS = (
    "exports by commodity",
    "export volume and value",
    "requirement 3.3",
)

HS_RE = re.compile(r"\((\d{4,4})\)")
# Filename year patterns we accept (in priority order):
#   "FY2021-2022"  -> fiscal year pair
#   "FY2021-22"    -> fiscal year pair short
#   "FY2021"       -> single fiscal year
#   "2018-2019"    -> calendar year pair (no FY prefix)
#   "2021"         -> single calendar year
FY_PATTERNS = [
    re.compile(r"FY(\d{4})[\s\-_/]+(\d{4})"),
    re.compile(r"FY(\d{4})[\s\-_/]+(\d{2})\b"),
    re.compile(r"FY(\d{4})"),
    re.compile(r"(?<!\d)(\d{4})[\s\-_/]+(\d{4})(?!\d)"),
    re.compile(r"(?<!\d)(\d{4})(?!\d)"),
]


# %% MARK: 3. Parse workbooks
def parse_fy(fname: str):
    """Return (label, start_year, end_year) from filename. None if no year detected."""
    for pat in FY_PATTERNS:
        m = pat.search(fname)
        if not m:
            continue
        groups = m.groups()
        a = int(groups[0])
        if a < 1990 or a > 2030:
            continue
        if len(groups) > 1 and groups[1]:
            b_raw = groups[1]
            b = int(b_raw) if len(b_raw) == 4 else int(str(a)[:2] + b_raw)
        else:
            b = a
        return m.group(0), a, b
    return None, None, None


def classify_section(text: str):
    if not text:
        return None
    t = text.lower()
    if any(h in t for h in PRODUCTION_HEADERS):
        return "production"
    if any(h in t for h in EXPORTS_HEADERS):
        return "exports"
    return None


def parse_commodity_label(label: str):
    """
    'Crude oil (2709), volume'   -> ('Crude oil', '2709', 'volume')
    'Coal (2701), value'         -> ('Coal',     '2701', 'value')
    """
    if not isinstance(label, str):
        return None, None, None
    s = label.strip().lstrip("'")
    m = HS_RE.search(s)
    hs = m.group(1) if m else None
    name = HS_RE.sub("", s)
    measure = None
    for kw in ("volume", "value"):
        if name.lower().rstrip().endswith(kw) or f", {kw}" in name.lower():
            measure = kw
            name = re.sub(rf"[,\s]+{kw}\s*$", "", name, flags=re.IGNORECASE)
            break
    return name.strip().rstrip(",").strip(), hs, measure


def _to_number(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if not s or s.lower().startswith(("not appli", "n/a", "na", "n.a.", "-")):
        return None
    s = s.replace(",", "").replace(" ", "")
    try:
        return float(s)
    except ValueError:
        return None


def parse_workbook(fp: Path, iso3: str):
    """Return list of dict rows for one XLSX."""
    fy_label, fy_start, fy_end = parse_fy(fp.name)
    out = []
    try:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    except Exception as exc:
        return [{"_error": f"{fp.name}: {exc}"}]

    target = None
    for cand in ("Part 2 - Disclosure checklist", "Part 2 - Disclosure  checklist"):
        if cand in wb.sheetnames:
            target = cand
            break
    if target is None:
        wb.close()
        return out

    ws = wb[target]
    current_section = None
    # Label can sit in column A or column B depending on the template version.
    # Detect dynamically: find a column index in [0,1,2] that holds a commodity
    # row, then keep using it for the rest of the sheet. Section headers ride
    # in the same column as the labels.
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue

        # Look for a section header in any of the first three columns
        for c in row[:3]:
            sec = classify_section(c if isinstance(c, str) else "")
            if sec is not None:
                current_section = sec
                break
        if current_section is None:
            continue

        # Pick the column that holds the label by finding the first cell that
        # parses as a commodity ("XX (HS), volume|value")
        label_col = None
        for j in (0, 1, 2):
            if j >= len(row):
                continue
            cell = row[j]
            if isinstance(cell, str) and HS_RE.search(cell):
                _, _hs, _meas = parse_commodity_label(cell)
                if _hs and _meas:
                    label_col = j
                    break
        if label_col is None:
            continue

        label = row[label_col]
        name, hs, measure = parse_commodity_label(label)
        if not measure or not hs:
            continue

        # Template uses empty spacer columns between label/value/unit/source.
        # Walk the row past the label and pick the first three non-None cells
        # as value, unit, source in that order.
        non_empty = []
        for k in range(label_col + 1, len(row)):
            if row[k] is None:
                continue
            non_empty.append(row[k])
            if len(non_empty) >= 3:
                break
        value_cell = non_empty[0] if len(non_empty) >= 1 else None
        unit_cell = non_empty[1] if len(non_empty) >= 2 else None
        src_cell = non_empty[2] if len(non_empty) >= 3 else None

        value = _to_number(value_cell)
        if value is None:
            continue

        out.append({
            "iso3": iso3,
            "fy_label": fy_label or "",
            "fy_start_year": fy_start or "",
            "fy_end_year": fy_end or "",
            "section": current_section,
            "commodity_name": name,
            "hs_code": hs,
            "category": HS_TO_CATEGORY.get(hs, "other"),
            "measure": measure,
            "value": value,
            "unit": str(unit_cell).strip() if unit_cell else "",
            "source_text": (str(src_cell)[:200] if src_cell else ""),
            "source_file": fp.name,
        })
    wb.close()
    return out


# %% MARK: 4. Run and write output
def main():
    if not EITI_ROOT.exists():
        print(f"ERROR: {EITI_ROOT} does not exist")
        return

    all_rows = []
    skipped = []
    for country_dir in sorted(EITI_ROOT.iterdir()):
        if not country_dir.is_dir():
            continue
        iso3 = NAME_TO_ISO3.get(country_dir.name)
        if not iso3:
            skipped.append(country_dir.name)
            continue
        xlsx_files = list(country_dir.glob("*.xlsx"))
        if not xlsx_files:
            continue
        print(f"  {iso3:3} {country_dir.name}: {len(xlsx_files)} xlsx files")
        for fp in xlsx_files:
            rows = parse_workbook(fp, iso3)
            errors = [r for r in rows if r.get("_error")]
            for e in errors:
                print(f"    ERROR {e['_error']}")
            rows = [r for r in rows if not r.get("_error")]
            all_rows.extend(rows)

    if skipped:
        print(f"\nSkipped (no iso3 mapping): {skipped}")

    if not all_rows:
        print("No data extracted.")
        return

    cols = [
        "iso3", "fy_label", "fy_start_year", "fy_end_year", "section",
        "commodity_name", "hs_code", "category", "measure", "value",
        "unit", "source_text", "source_file",
    ]
    with open(OUT_PATH, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for r in all_rows:
            w.writerow(r)
    print(f"\nWrote {len(all_rows):,} rows -> {OUT_PATH}")

    # Quick coverage stats
    by_iso = {}
    for r in all_rows:
        by_iso.setdefault(r["iso3"], set()).add(r["fy_end_year"])
    print("\nFiscal-end-year coverage per country:")
    for iso, yrs in sorted(by_iso.items()):
        years = sorted(y for y in yrs if y)
        print(f"  {iso}: {years}")


if __name__ == "__main__":
    main()
