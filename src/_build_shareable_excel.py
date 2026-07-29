"""
Build a plain-language Excel of the baseline unitary-taxation results for
sharing (the numbers behind the paper's Table 2 — change in taxable profits by
formula — and Table 3 — change in tax revenue by formula, reported sample).

Reads the two table CSVs written by 7b_formula_results.py and writes a
formatted workbook with a Read-me sheet, one sheet per table, income-group
subtotal rows in bold, and a country-name index. TJN brand colours; numbers as
US$ million per year (yearly average 2016–2022 excl. 2020, constant 2025 USD).

Output: output/app/unitary_taxation_baseline_results.xlsx
Refresh after a pipeline rerun: `python _build_shareable_excel.py`.
Not part of the replication package (dissemination deliverable).
"""
import os
import sys
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

_ROOT = Path(__file__).resolve().parent.parent
TABLES = _ROOT / "output" / "paper" / "main_text" / "tables"
OUT = _ROOT / "output" / "app" / "unitary_taxation_baseline_results.xlsx"

# TJN brand
GOLD = "FFD371"
TEAL = "45636C"
GREEN = "50805E"
INK = "26312F"
BAND = "F3F0E6"
WHITE = "FFFFFF"

# (sheet title, csv, heading, blurb, super-header over the formula block,
#  renamed reference column — moved LAST, as in the paper tables)
SHEETS = [
    ("Taxable profits", "table2_taxable_profits_by_formula__reported_only.csv",
     "Change in taxable profits by apportionment formula",
     "How much taxable profit each country would gain (+) or lose (−) if "
     "multinationals' profits were reallocated to where their real activity is.",
     "Annual change in taxable profits under unitary taxation",
     "Currently reported profits"),
    ("Tax revenue", "table3_tax_revenue_by_formula__reported_only.csv",
     "Change in tax revenue by apportionment formula",
     "The resulting change in each country's corporate tax revenue, valuing "
     "reallocated profit at statutory rates for gains and effective rates for losses.",
     "Annual change in corporate income tax revenue under unitary taxation",
     "Currently collected tax revenue"),
]

ESWATINI_NOTE = (
    "Note on Eswatini: Eswatini's projected loss is driven by the exceptionally high profits "
    "US-headquartered multinationals report there — 94% of all multinational profit reported in "
    "Eswatini in 2022 — a pattern consistent with Coca-Cola's concentrate operation in the country. "
    "That operation has been documented as benefiting from a corporate tax rate of about 6% and "
    "financial secrecy (Sharife 2015, 100Reporters) and lies at the centre of Coca-Cola's "
    "transfer-pricing dispute with the US tax authorities (Financial Times, Foley and Meyer 2024).")
FLAG_NOTES = (
    "[1] Percentage not calculated: the country's multinationals reported only losses, so the "
    "positive-profits base is zero.   [2] after a country name: thin data coverage — the estimate "
    "rests on few reported cells, parents or years; interpret with caution.   [3] after a "
    "percentage: extreme value (above 1,000%) — the country's net reported base is close to zero, "
    "so the ratio is denominator-driven; read the absolute change instead.")

_thin = Side(style="thin", color="D9D5C8")
BORDER = Border(bottom=_thin)


def _is_heading(name):
    s = str(name).strip()
    return s == s.upper() and len(s) > 3 and any(ch.isalpha() for ch in s)


# Display labels for the group-subtotal rows (the source CSVs carry them in
# all-caps; bold + shading already set them apart, so no caps in the workbook).
GROUP_LABELS = {
    "LOW-INCOME": "Low-income countries",
    "LOWER-MIDDLE-INCOME": "Lower-middle-income countries",
    "UPPER-MIDDLE-INCOME": "Upper-middle-income countries",
    "HIGH-INCOME": "High-income countries",
    "TAX HAVENS": "Tax havens",
}


def _write_sheet(wb, title, csv_name, heading, blurb, super_header, ref_name):
    p = TABLES / csv_name
    if not p.exists():
        print(f"  [skip] {csv_name} not found — run 7b_formula_results.py first")
        return False
    df = pd.read_csv(p)
    # ---- mirror the paper tables: reference column LAST, short % headers ----
    ref = df.columns[1]                              # after Country
    df = df[[df.columns[0]] + list(df.columns[2:]) + [ref]].rename(columns={ref: ref_name})
    ncol = len(df.columns)
    formulas = [c for c in df.columns[1:-1] if "(%" not in c]   # the 4 formula names
    ws = wb.create_sheet(title)

    # title rows
    ws["A1"] = heading
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=INK)
    ws["A2"] = blurb
    ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="5B6B67")
    ws["A3"] = ("US$ million per year · yearly average 2016–2022 (2020 excluded) · "
                "constant 2025 USD · directly reported OECD country-by-country data")
    ws["A3"].font = Font(name="Calibri", size=9, color="93A19C")

    # ---- three header rows, as in the paper tables ----
    # row 5: super-header over the formula block | reference label (merged down)
    # row 6: formula names, each spanning its US$ m + % pair
    # row 7: US$ m / % units ("US$ m" under the reference too)
    top, mid, unit = 5, 6, 7
    def _hcell(r, jcol, val, size=10):
        c = ws.cell(row=r, column=jcol, value=val)
        c.font = Font(name="Calibri", size=size, bold=True, color=INK)
        c.fill = PatternFill("solid", fgColor=GOLD)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
        return c
    for r in (top, mid, unit):                      # gold band everywhere first
        for j in range(1, ncol + 1):
            _hcell(r, j, None)
    ws.cell(row=top, column=1, value="Country")
    ws.merge_cells(start_row=top, start_column=1, end_row=unit, end_column=1)
    ws.cell(row=top, column=1).alignment = Alignment(horizontal="left", vertical="center")
    _hcell(top, 2, super_header, size=11)
    ws.merge_cells(start_row=top, start_column=2, end_row=top, end_column=ncol - 1)
    for k, f in enumerate(formulas):
        j = 2 + 2 * k
        _hcell(mid, j, f)
        ws.merge_cells(start_row=mid, start_column=j, end_row=mid, end_column=j + 1)
        _hcell(unit, j, "US$ m"); _hcell(unit, j + 1, "%")
    _hcell(top, ncol, ref_name)
    ws.merge_cells(start_row=top, start_column=ncol, end_row=mid, end_column=ncol)
    _hcell(unit, ncol, "US$ m")

    # body — money columns as whole millions, % columns with one decimal
    pct_col = {j for j, c in enumerate(df.columns, start=1) if "(%" in c}
    hdr_row = unit
    for i, (_, row) in enumerate(df.iterrows(), start=hdr_row + 1):
        heading_row = _is_heading(row.iloc[0])
        for j, val in enumerate(row, start=1):
            if j == 1 and heading_row:
                val = GROUP_LABELS.get(str(val).strip(), str(val).strip().capitalize())
            c = ws.cell(row=i, column=j)
            if isinstance(val, str):                # flag-bearing columns are text;
                try:                                 # store plain numbers as numbers
                    val = float(val.replace(",", ""))
                except ValueError:
                    pass
            if isinstance(val, float) and pd.notna(val):
                if j in pct_col:
                    c.value = round(val, 1); c.number_format = '#,##0.0'
                else:
                    c.value = round(val);    c.number_format = '#,##0'
            else:
                c.value = "" if (isinstance(val, float) and pd.isna(val)) else val
            c.alignment = Alignment(horizontal="left" if j == 1 else "right")
            if heading_row:
                c.font = Font(name="Calibri", size=10, bold=True, color=INK)
                c.fill = PatternFill("solid", fgColor=BAND)
            else:
                c.font = Font(name="Calibri", size=10, color=INK)

    # notes under the table (flags + Eswatini), as in the paper
    note_row = hdr_row + len(df) + 2
    for txt in (FLAG_NOTES, ESWATINI_NOTE):
        c = ws.cell(row=note_row, column=1, value=txt)
        c.font = Font(name="Calibri", size=9, italic=True, color="5B6B67")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=note_row, start_column=1,
                       end_row=note_row, end_column=ncol)
        ws.row_dimensions[note_row].height = 42
        note_row += 2

    # widths + freeze
    ws.column_dimensions["A"].width = 30
    for j in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(j)].width = 13
    ws.column_dimensions[get_column_letter(ncol)].width = 15
    ws.freeze_panes = f"B{hdr_row + 1}"
    ws.sheet_view.showGridLines = False
    print(f"  wrote sheet '{title}' ({len(df)} rows)")
    return True


def _readme(wb):
    ws = wb.create_sheet("Read me", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    lines = [
        ("Unitary taxation — baseline results", 16, True, INK),
        ("Tax Justice Network", 11, False, TEAL),
        ("", 10, False, INK),
        ("What this shows", 12, True, INK),
        ("If multinational groups were taxed as single firms — their global profit shared "
         "out among countries in proportion to where their real activity (employees, sales, "
         "assets) actually is — how would each country's taxable profits and corporate tax "
         "revenue change? These are the paper's headline results.", 10, False, INK),
        ("", 10, False, INK),
        ("The two sheets", 12, True, INK),
        ("• Taxable profits — the change in each country's taxable profit base, by formula.", 10, False, INK),
        ("• Tax revenue — the resulting change in corporate tax revenue, by formula.", 10, False, INK),
        ("Each is shown for four apportionment formulas; 'Sales & employees' is the headline. "
         "The '(%)' columns express the change relative to the country's own current figures "
         "in the same data (see below).", 10, False, INK),
        ("", 10, False, INK),
        ("How to read the numbers", 12, True, INK),
        ("• Units: US$ million per year, averaged over 2016–2022 with 2020 excluded, in "
         "constant 2025 US dollars.", 10, False, INK),
        ("• A positive number means the country gains taxable profit / tax revenue; negative "
         "means it loses.", 10, False, INK),
        ("• Percentages compare with the country's own reported figures in the same OECD "
         "country-by-country data — the taxable-profit change against current positive "
         "reported profits, the tax-revenue change against the corporate income tax the "
         "sampled multinationals report paying there. They are not relative to a country's "
         "official total corporate tax revenue.", 10, False, INK),
        ("• Bold shaded rows are income-group subtotals; the countries below each belong to it.", 10, False, INK),
        ("", 10, False, INK),
        ("Coverage", 12, True, INK),
        ("Multinational groups with over €750 million in annual revenue that report country "
         "by country to tax authorities — roughly three-quarters of global multinational "
         "profit. Figures are not scaled up to the full population (the paper reports that "
         "scale-up separately). Directly reported data only, no imputation.", 10, False, INK),
        ("", 10, False, INK),
        ("Full method: the methodology note accompanying the paper. "
         "Estimates are research results, not forecasts.", 9, False, "93A19C"),
    ]
    for i, (text, size, bold, color) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        c.font = Font(name="Calibri", size=size, bold=bold, color=color)
        c.alignment = Alignment(wrap_text=True, vertical="top")


# Shareable copy for the team ("final" deliverables folder on SharePoint).
FINAL_DIR = Path(r"C:\Users\aliso\Tax Justice Network Ltd\TJN - Shared Documents"
                 r"\Workstreams\Scale of Tax Injustice\Unitary taxation\final")


def main():
    wb = Workbook()
    wb.remove(wb.active)   # drop default sheet
    any_written = False
    for title, csv_name, heading, blurb, super_header, ref_name in SHEETS:
        any_written |= _write_sheet(wb, title, csv_name, heading, blurb,
                                    super_header, ref_name)
    if not any_written:
        raise SystemExit("No source tables found — run 7b_formula_results.py first.")
    _readme(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    try:
        import shutil
        FINAL_DIR.mkdir(parents=True, exist_ok=True)
        shutil.copy2(OUT, FINAL_DIR / "Unitary_taxation_results.xlsx")
        print(f"copied to {FINAL_DIR / 'Unitary_taxation_results.xlsx'}")
    except Exception as e:
        print(f"[warn] could not copy to final folder: {e}")


if __name__ == "__main__":
    main()
