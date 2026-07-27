# %%
"""
Per-country resource-correction table: the pre-profit and post-profit resource
payments applied in the extractive correction, with provenance (why the value is
what it is) and a plausibility sanity check.

Columns (per SOURCE country = iso_partner, summed 2016–2022, USD m):
  pre_profit_payments      royalties / production-sharing / fees deducted BEFORE
                           CbCR profit (added back in the incl_resource view)
  post_profit_payments     resource CIT / special taxes paid FROM profit
  equity_income            SOE / state-participation income
  actual_resource_contribution = pre + post + equity (total state take)
  resource_profit_base     the profit pool stripped out in the excl_resource view
  reported_profit / reported_tax        for scale
  contribution_pct_profit  actual_resource_contribution / |reported_profit|
  src_*                    share of the payments coming from each provenance
                           (EITI actual data / GRD / manual / rent-proxy estimate)
  main_commodity           dominant commodity in the resource panel
  sanity_flag              heuristic plausibility flags

Reads data/final/cbcr_main_incl_resource.csv + _excl_resource.csv (the applied
correction) and data/intermediate/extractive/resource_payments_by_hq_source_yearly.csv
(provenance). Writes output/unitary_taxation/across_samples/resource_correction/tables/
(via the output_dirs remap of the "deliverables/resource_correction" topic).

Usage: python 9f_resource_correction_table.py
"""
import os
import subprocess
import sys
import tempfile

import numpy as np
import pandas as pd

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except AttributeError:
    pass

from config import output_dirs, data_final, data_intermediate_extractive

YEARS = list(range(2016, 2023))

# Plain-language explanation written as a SECOND header line (row directly below
# the column names) in the output table. src_* columns are described generically.
COLUMN_DESCRIPTIONS = {
    "iso_partner": "Source country ISO3 code (where the resource is extracted)",
    "partner_jurisdiction": "Source country name",
    "wb_income_group": "World Bank income group (investment_hub = recognised tax haven)",
    "main_commodity": "Dominant commodity in this country's resource-payment panel",
    "pre_profit_payments_usd_m": "Royalties / production-sharing / fees deducted BEFORE CbCR "
        "profit, USD millions, 2016-2022 total (added back in the incl_resource view)",
    "post_profit_payments_usd_m": "Resource corporate income tax / special taxes paid FROM "
        "profit, USD millions, 2016-2022 total",
    "equity_income_usd_m": "State-owned-enterprise / state-participation income, USD millions, "
        "2016-2022 total",
    "actual_resource_contribution_usd_m": "Total state take = pre-profit + post-profit + equity, "
        "USD millions, 2016-2022 total",
    "resource_profit_base_usd_m": "Profit pool removed from the country in the excl_resource "
        "view, USD millions, 2016-2022 total",
    "reported_profit_usd_m": "Country's reported profit before income tax (for scale), USD "
        "millions, 2016-2022 total",
    "reported_tax_usd_m": "Country's reported corporate income tax paid (for scale), USD "
        "millions, 2016-2022 total",
    "contribution_pct_profit": "Total state take as a percentage of the country's absolute "
        "reported profit",
    "contribution_pct_gdp_per_yr": "Annual state take as a percentage of GDP (total take / 7 "
        "years / mean annual GDP)",
    "sanity_flag": "Plausibility flags: ok / no_resource_correction / >50%GDP-per-yr check / "
        "mostly_rent_proxy (estimated) / negative_component check",
}
_SRC_NAMES = {"eiti": "EITI-reported company payments", "grd": "ICTD Government Revenue Dataset",
              "manual": "manual desk research", "rent_proxy": "rent-proxy estimate"}


def _describe(col):
    if col in COLUMN_DESCRIPTIONS:
        return COLUMN_DESCRIPTIONS[col]
    if col.startswith("src_"):
        return f"Share of the total state take sourced from {_SRC_NAMES.get(col[4:], col[4:])}"
    return ""


def rd(path, **kw):
    try:
        return pd.read_csv(path, **kw)
    except (FileNotFoundError, OSError):
        dst = os.path.join(tempfile.gettempdir(), "rct_" + os.path.basename(path))
        for cp in (r"C:\Program Files\Git\usr\bin\cp.exe", "cp"):
            try:
                subprocess.run([cp, os.fspath(path), dst], check=True)
                return pd.read_csv(dst, **kw)
            except Exception:
                continue
        raise


def main():
    tables_dir, _ = output_dirs("deliverables/resource_correction")

    incl = rd(f"{data_final}cbcr_main_incl_resource.csv", low_memory=False)
    incl = incl[incl["year"].isin(YEARS)]
    g = incl.groupby("iso_partner", as_index=False).agg(
        pre_profit_payments_usd=("pre_profit_payments_usd", "sum"),
        post_profit_payments_usd=("post_profit_payments_usd", "sum"),
        equity_income_usd=("equity_income_usd", "sum"),
        actual_resource_contribution_usd=("actual_resource_contribution_usd", "sum"),
        reported_profit_usd=("profit_loss_before_income_tax_corrected", "sum"),
        reported_tax_usd=("income_tax_paid_on_cash_basis", "sum"),
        partner_jurisdiction=("partner_jurisdiction", "first"),
        wb_income_group=("wb_income_group", "first"),
    )
    # GDP (for the sanity ratio) from the main cleaned file.
    gpop = rd(f"{data_final}cbcr_main_allsubgroupsonly.csv",
              usecols=["iso_partner", "year", "gdp_current_usd"])
    gpop = gpop[gpop["year"].isin(YEARS)].drop_duplicates(["iso_partner", "year"])
    gdp = gpop.groupby("iso_partner", as_index=False)["gdp_current_usd"].mean().rename(
        columns={"gdp_current_usd": "gdp_usd"})
    g = g.merge(gdp, on="iso_partner", how="left")

    excl = rd(f"{data_final}cbcr_main_excl_resource.csv",
              usecols=["iso_partner", "year", "resource_profit_base_usd"])
    excl = excl[excl["year"].isin(YEARS)]
    rpb = excl.groupby("iso_partner", as_index=False)["resource_profit_base_usd"].sum()
    g = g.merge(rpb, on="iso_partner", how="left")

    # Provenance + commodity from the resource panel (by source country).
    panel = rd(f"{data_intermediate_extractive}resource_payments_by_hq_source_yearly.csv")
    panel = panel[panel["year"].isin(YEARS)].copy()
    panel["tot"] = (panel["pre_profit_payments_usd"].fillna(0)
                    + panel["post_profit_payments_usd"].fillna(0)
                    + panel["equity_income_usd"].fillna(0))
    prov = panel.groupby(["source_iso3", "data_source"])["tot"].sum().unstack(fill_value=0)
    prov_share = prov.div(prov.sum(axis=1).replace(0, np.nan), axis=0)
    prov_share.columns = ["src_" + c for c in prov_share.columns]
    prov_share = prov_share.reset_index().rename(columns={"source_iso3": "iso_partner"})
    g = g.merge(prov_share, on="iso_partner", how="left")
    commodity = (panel.groupby(["source_iso3", "commodity"])["tot"].sum()
                 .reset_index().sort_values("tot", ascending=False)
                 .drop_duplicates("source_iso3")
                 .rename(columns={"source_iso3": "iso_partner", "commodity": "main_commodity"}))
    g = g.merge(commodity[["iso_partner", "main_commodity"]], on="iso_partner", how="left")

    # Derived + sanity.
    # contribution_pct_profit: both numerator and reported profit are 7-year sums
    # → consistent. contribution_pct_gdp: GDP is a MEAN ANNUAL value, so annualise
    # the contribution (÷ number of years) before dividing, otherwise a multi-year
    # sum vs one year's GDP inflates the ratio ~7×.
    n_years = len(YEARS)
    g["contribution_pct_profit"] = (
        100 * g["actual_resource_contribution_usd"] / g["reported_profit_usd"].abs()
    ).replace([np.inf, -np.inf], np.nan)
    g["contribution_pct_gdp_per_yr"] = (
        100 * (g["actual_resource_contribution_usd"] / n_years) / g["gdp_usd"]
    ).replace([np.inf, -np.inf], np.nan)

    def flag(r):
        f = []
        if r["actual_resource_contribution_usd"] <= 0:
            f.append("no_resource_correction")
            return ";".join(f)
        if pd.notna(r["contribution_pct_gdp_per_yr"]) and r["contribution_pct_gdp_per_yr"] > 50:
            f.append("contribution>50%GDP/yr_check")
        if r.get("src_rent_proxy", 0) and r["src_rent_proxy"] > 0.66:
            f.append("mostly_rent_proxy(estimated)")
        if r["post_profit_payments_usd"] < 0 or r["pre_profit_payments_usd"] < 0:
            f.append("negative_component_check")
        return ";".join(f) if f else "ok"

    g["sanity_flag"] = g.apply(flag, axis=1)

    # Scale to USD millions for readability.
    musd = ["pre_profit_payments_usd", "post_profit_payments_usd", "equity_income_usd",
            "actual_resource_contribution_usd", "resource_profit_base_usd",
            "reported_profit_usd", "reported_tax_usd"]
    for c in musd:
        g[c + "_m"] = g[c] / 1e6

    keep = (["iso_partner", "partner_jurisdiction", "wb_income_group", "main_commodity"]
            + [c + "_m" for c in musd]
            + ["contribution_pct_profit", "contribution_pct_gdp_per_yr"]
            + [c for c in g.columns if c.startswith("src_")]
            + ["sanity_flag"])
    out = g[keep].sort_values("actual_resource_contribution_usd_m", ascending=False)

    # Two-row header: column name on the first line, a plain-language explanation
    # on the second line directly below it, with the data kept NUMERIC.
    descriptions = [_describe(c) for c in out.columns]

    # CSV: MultiIndex column header writes the two header lines, data stays numeric.
    csv_path = tables_dir / "resource_correction_by_country_2016_22.csv"
    out_csv = out.copy()
    out_csv.columns = pd.MultiIndex.from_tuples(
        list(zip(out.columns, descriptions)), names=["column", "description"])
    out_csv.to_csv(csv_path, index=False)
    print(f"wrote {csv_path} ({len(out)} countries with rows; "
          f"{int((g['actual_resource_contribution_usd'] > 0).sum())} with a correction)")

    # Excel: single header via pandas (numeric), then insert the description row
    # directly beneath it with openpyxl (pandas can't write MultiIndex cols + index=False).
    xlsx_path = tables_dir / "resource_correction_by_country_2016_22.xlsx"
    try:
        out.to_excel(xlsx_path, index=False)
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active
        ws.insert_rows(2)
        for j, d in enumerate(descriptions, start=1):
            ws.cell(row=2, column=j, value=d)
        wb.save(xlsx_path)
        print(f"wrote {xlsx_path}")
    except Exception as e:
        print(f"  (xlsx skipped: {e})")

    # Headline print: top resource-contribution countries.
    top = out[out["actual_resource_contribution_usd_m"] > 0].head(25)
    print("\nTop resource-correction countries (2016–2022, USD m):")
    print(f"  {'iso':4s} {'country':24s} {'pre':>9s} {'post':>9s} {'equity':>8s} "
          f"{'total':>9s} {'%profit':>7s} {'commodity':>9s}  flag")
    for _, r in top.iterrows():
        print(f"  {r['iso_partner']:4s} {str(r['partner_jurisdiction'])[:24]:24s} "
              f"{r['pre_profit_payments_usd_m']:>9.0f} {r['post_profit_payments_usd_m']:>9.0f} "
              f"{r['equity_income_usd_m']:>8.0f} {r['actual_resource_contribution_usd_m']:>9.0f} "
              f"{(r['contribution_pct_profit'] if pd.notna(r['contribution_pct_profit']) else float('nan')):>7.0f} "
              f"{str(r['main_commodity'])[:9]:>9s}  {r['sanity_flag']}")


if __name__ == "__main__":
    main()
